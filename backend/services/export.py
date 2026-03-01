import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

BULAN = ["","Januari","Februari","Maret","April","Mei","Juni",
         "Juli","Agustus","September","Oktober","November","Desember"]


def generate_excel(data: list[dict], bulan: int, tahun: int) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Absensi"

    thin   = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    # Title
    ws.merge_cells("A1:K1")
    ws["A1"] = f"LAPORAN REKAP ABSENSI — {BULAN[bulan].upper()} {tahun}"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="0F172A")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    # Header
    headers    = ["No","ID","Nama Karyawan","Divisi","Hari Kerja",
                  "Hadir","Terlambat","Absen","Izin","Total Jam","% Hadir"]
    col_widths = [5, 12, 28, 18, 12, 8, 10, 8, 8, 16, 12]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font      = Font(bold=True, color="E2E8F0", size=9)
        cell.fill      = PatternFill("solid", fgColor="334155")
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[2].height = 22

    # Data
    for i, row in enumerate(data, 1):
        r   = i + 2
        pct = float(row.get("persen_kehadiran", 0))
        pct_color = "059669" if pct >= 90 else ("D97706" if pct >= 75 else "DC2626")

        values = [
            i,
            row.get("kode_karyawan",""),
            row.get("nama",""),
            row.get("divisi",""),
            row.get("total_hari", 0),
            row.get("hari_hadir", 0),
            row.get("hari_terlambat", 0),
            row.get("hari_absen", 0),
            row.get("hari_izin", 0),
            _fmt_menit(int(row.get("total_menit_kerja", 0))),
            f"{pct:.1f}%",
        ]
        colors = [None,None,None,None,None,
                  "059669","D97706","DC2626","3B82F6",None,pct_color]

        for c, (v, color) in enumerate(zip(values, colors), 1):
            cell            = ws.cell(row=r, column=c, value=v)
            cell.border     = border
            cell.alignment  = left if c == 3 else center
            if color:
                cell.font   = Font(color=color, bold=(c in (6,7,8,9,11)), size=10)
        if i % 2 == 0:
            for c in range(1, 12):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _fmt_menit(m: int) -> str:
    if not m: return "—"
    j, s = divmod(m, 60)
    if not j: return f"{s}m"
    if not s: return f"{j}j"
    return f"{j}j {s}m"